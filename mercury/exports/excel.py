"""Weekly spreadsheets, rendered on demand from the database with dynamic charge codes."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import Config
from ..models import list_custom_items, list_jobs, week_label, week_summary
from ..rates import get_item_list, rate_label

NAVY = "0F172A"
SLATE = "1E293B"
ACCENT = "22D3EE"
LIGHT = "F1F5F9"

_title_font = Font(size=16, bold=True, color="FFFFFF")
_header_font = Font(bold=True, size=10, color="FFFFFF")
_label_font = Font(bold=True, size=10)
_money_fmt = '"$"#,##0.00'

_thin = Side(style="thin", color="CBD5E1")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header_row(ws, row: int, ncols: int, fill: str = SLATE) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border


def _build(start: date, end: date, include_pay: bool) -> openpyxl.Workbook:
    summary = week_summary(start, end)
    item_list = get_item_list()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Installations"

    headers = ["Date", "Address", "Order Number"] + item_list + ["Notes"]
    if include_pay:
        headers.append("Job Total")
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1, value="Mercury Installation Tracker")
    title.font = _title_font
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=3, column=1, value="Technician:").font = _label_font
    ws.cell(row=3, column=2, value=Config.TECH_NAME)
    ws.cell(row=4, column=1, value="Week:").font = _label_font
    ws.cell(row=4, column=2,
            value=f"{start.strftime('%m/%d/%Y')} - {end.strftime('%m/%d/%Y')}")
    ws.cell(row=3, column=4, value="Generated:").font = _label_font
    ws.cell(row=3, column=5, value=datetime.now().strftime("%m/%d/%Y %I:%M %p"))

    for col, name in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=name)
    _style_header_row(ws, 5, ncols)
    ws.freeze_panes = "A6"

    row_num = 6
    for job in sorted(summary["jobs"], key=lambda j: (j["work_date"], j["created_at"])):
        ws.cell(row=row_num, column=1, value=job["work_date"])
        ws.cell(row=row_num, column=2, value=job["address"])
        ws.cell(row=row_num, column=3, value=job["order_number"])
        for idx, item in enumerate(item_list):
            qty = job["items"].get(item)
            ws.cell(row=row_num, column=idx + 4, value=qty if qty else None)
        ws.cell(row=row_num, column=len(item_list) + 4, value=job["notes"] or "")
        if include_pay:
            cell = ws.cell(row=row_num, column=ncols, value=job["total"])
            cell.number_format = _money_fmt
        for col in range(1, ncols + 1):
            ws.cell(row=row_num, column=col).border = _border
        row_num += 1

    if row_num > 6:
        total_row = row_num
        ws.cell(row=total_row, column=1, value="TOTALS").font = _label_font
        for idx, item in enumerate(item_list):
            col = idx + 4
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f"=SUM({letter}6:{letter}{total_row - 1})")
            cell.font = _label_font
        if include_pay:
            letter = get_column_letter(ncols)
            cell = ws.cell(row=total_row, column=ncols,
                           value=f"=SUM({letter}6:{letter}{total_row - 1})")
            cell.font = _label_font
            cell.number_format = _money_fmt
        for col in range(1, ncols + 1):
            c = ws.cell(row=total_row, column=col)
            c.fill = PatternFill("solid", fgColor="E2E8F0")
            c.border = _border

    widths = [12, 34, 16] + [16] * len(item_list) + [40]
    if include_pay:
        widths.append(14)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    _add_custom_sheet(wb, start, end, include_pay)
    if include_pay:
        _add_summary_sheet(wb, summary, item_list)
    return wb


def _add_custom_sheet(wb, start: date, end: date, include_pay: bool) -> None:
    items = list_custom_items(start, end)
    ws = wb.create_sheet("Custom Items")
    headers = ["Date", "Item Name", "Description", "Qty", "Rate", "Total", "Bill To"]
    if not include_pay:
        headers = headers[:4] + ["Bill To"]
    for col, name in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=name)
    _style_header_row(ws, 1, len(headers))

    for r, item in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=item["work_date"])
        ws.cell(row=r, column=2, value=item["name"])
        ws.cell(row=r, column=3, value=item["description"])
        ws.cell(row=r, column=4, value=item["qty"])
        if include_pay:
            ws.cell(row=r, column=5, value=item["rate"]).number_format = _money_fmt
            ws.cell(row=r, column=6, value=item["total"]).number_format = _money_fmt
            ws.cell(row=r, column=7, value=item["bill_to"].upper())
        else:
            ws.cell(row=r, column=5, value=item["bill_to"].upper())

    for i, width in enumerate([12, 28, 40, 10, 12, 14, 12][:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _add_summary_sheet(wb, summary: dict, item_list: list[str]) -> None:
    ws = wb.create_sheet("Pay Summary")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    ws.merge_cells("A1:D1")
    banner = ws.cell(row=1, column=1, value=f"Pay Summary · {summary['range_label']}")
    banner.font = _title_font
    banner.fill = PatternFill("solid", fgColor=NAVY)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, name in enumerate(["Item", "Qty", "Rate", "Pay"], 1):
        ws.cell(row=3, column=col, value=name)
    _style_header_row(ws, 3, 4)

    row = 4
    for item in item_list:
        qty = summary["task_counts"].get(item, 0)
        if not qty:
            continue
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=rate_label(item))
        ws.cell(row=row, column=4,
                value=summary["task_pay"].get(item, 0.0)).number_format = _money_fmt
        row += 1

    row += 1
    for label, value in (
        ("Jobs subtotal", summary["jobs_total"]),
        ("Custom items subtotal", summary["custom_total"]),
        ("WEEK TOTAL", summary["weekly"]),
        ("Jobs logged", summary["job_count"]),
        ("Days worked", summary["days_worked"]),
        ("Average per job", summary["avg_per_job"]),
        ("Average per day", summary["avg_per_day"]),
    ):
        ws.cell(row=row, column=1, value=label).font = _label_font
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = _money_fmt
        if label == "WEEK TOTAL":
            cell.font = Font(bold=True, size=12)
        row += 1


def _write(wb, filename: str) -> Path:
    Config.ensure_dirs()
    path = Config.EXPORT_DIR / filename
    wb.save(path)
    return path


def build_contractor_workbook(start: date, end: date) -> Path:
    return _write(_build(start, end, include_pay=False),
                  f"Mercury Install Tracker {week_label(start, end)}.xlsx")


def build_personal_workbook(start: date, end: date) -> Path:
    return _write(_build(start, end, include_pay=True),
                  f"Personal Pay Report {week_label(start, end)}.xlsx")
